import os
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def process_player_data(player_data_folder, excel_file):
    # Create a new workbook
    workbook = Workbook()

    # First sheet for game data
    data_sheet = workbook.active
    data_sheet.title = "Spieldaten"

    # Add headers
    data_headers = [
        'Region',
        'Spieler',
        'Platzierung',
        'Spielmodus',
        'Dauer',
        'Alter',
        'Avatar-Level',
        'Units',
        'Traits'
    ]
    data_sheet.append(data_headers)

    # Second sheet for Units and Items
    unit_sheet = workbook.create_sheet(title="Units und Items")
    unit_headers = [
        'Region',
        'Spieler',
        'Einheit',
        'Platzierung',
        'Items'
    ]
    unit_sheet.append(unit_headers)

    # Iterate through each region folder
    for region_folder_name in os.listdir(player_data_folder):
        region_folder_path = os.path.join(player_data_folder, region_folder_name)

        # Skip files that are not folders
        if not os.path.isdir(region_folder_path):
            continue

        # Iterate through each player folder in the region
        for player_folder_name in os.listdir(region_folder_path):
            player_folder_path = os.path.join(region_folder_path, player_folder_name)

            # Skip files that are not folders
            if not os.path.isdir(player_folder_path):
                continue

            # Iterate through each JSON file in the player folder
            for file_name in os.listdir(player_folder_path):
                file_path = os.path.join(player_folder_path, file_name)

                # Skip files that are not JSON files
                if not file_name.endswith('.json'):
                    continue

                # Read JSON data from the file
                with open(file_path, 'r') as json_file:
                    data_list = json.load(json_file)

                # Check if the data is a list
                if not isinstance(data_list, list):
                    continue

                # Extract game data from the JSON objects
                for data in data_list:
                    # Write game data to the "Spieldaten" sheet
                    units = data.get('units', [])
                    unit_names = ', '.join(unit.get('unitName', '').replace("â€™", "'") for unit in units)
                    data_row = [
                        region_folder_name,
                        player_folder_name,
                        data.get('placement', ''),
                        data.get('gameMode', ''),
                        data.get('length', ''),
                        data.get('age', ''),
                        data.get('avatarLevel', ''),
                        unit_names,
                        ', '.join(data.get('traits', []))
                    ]
                    data_sheet.append(data_row)

                    # Extract Unit and Item data and write to the "Units und Items" sheet
                    for unit in units:
                        unit_row = [
                            region_folder_name,
                            player_folder_name,
                            unit.get('unitName', '').replace("â€™", "'"),
                            data.get('placement', ''),  # Add placement of the unit in the game
                            ', '.join(unit.get('items', []))
                        ]
                        unit_sheet.append(unit_row)

    # Set optimal column width for the "Spieldaten" sheet
    set_column_width(data_sheet)

    # Set optimal column width for the "Units und Items" sheet
    set_column_width(unit_sheet)

    # Save the Excel file
    workbook.save(excel_file)


def set_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width


# Adjust paths
player_data_folder = "D:\\TFTScraping\\playerlists"
excel_file = "D:\\TFTScraping\\tft_data.xlsx"

# Process data and create Excel file
process_player_data(player_data_folder, excel_file)