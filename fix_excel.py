import pandas as pd
from openpyxl import load_workbook

def find_and_replace(file_path, old_value, new_value):
    # Load the Excel file
    book = load_workbook(file_path)

    # Check if the sheet already exists
    if "Units und Items" in book.sheetnames:
        # Get the sheet by name
        sheet = book["Units und Items"]

        # Perform the search and replace in the "Items" column
        for cell in sheet["E"]:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace(old_value, new_value)

        # Save the modified Excel file
        book.save(file_path)
    else:
        print("Sheet 'Units und Items' does not exist in the Excel file.")

# Adjust the paths and values as needed
excel_file = "D:/TFTScraping/tft_data.xlsx"
old_value = "Titanâ€™s Resolve"
new_value = "Titan's Resolve"

# Perform the search and replace
find_and_replace(excel_file, old_value, new_value)
